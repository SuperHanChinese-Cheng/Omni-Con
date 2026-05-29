# Copyright (c) 2026 Chenglin Qiu (SHC - Super Han Chinese). All rights reserved.
# Licensed under the OmniCon Proprietary License. See LICENSE for details.
"""Table extraction engine — extracts tables from PDFs into Excel spreadsheets."""

import logging
from pathlib import Path

import openpyxl
import pdfplumber

from omnicon.core.job import ConversionJob
from omnicon.engines.base import BaseEngine, EngineError, UnsupportedRouteError

logger = logging.getLogger(__name__)


class TableEngine(BaseEngine):
    """Extracts tables from PDF files and writes them to XLSX using pdfplumber + openpyxl.

    Each page that contains tables gets its own worksheet (named "Page 1", "Page 2", etc.).
    If no tables are found in the entire PDF, the workbook is saved with a single sheet
    containing a note that no tables were detected.
    """

    priority = 15

    SUPPORTED_ROUTES: set[tuple[str, str]] = {
        ("pdf", "xlsx"),
    }

    def can_convert(self, src_fmt: str, dst_fmt: str) -> bool:
        """Check if this engine supports the given conversion route."""
        return (src_fmt.lower(), dst_fmt.lower()) in self.SUPPORTED_ROUTES

    def convert(self, job: ConversionJob) -> Path:
        """Execute PDF-to-XLSX table extraction. Thread-safe (no shared mutable state)."""
        route = (job.src_format, job.output_format.lower())
        if route != ("pdf", "xlsx"):
            raise UnsupportedRouteError(
                f"Route {route} not supported by {self.__class__.__name__}"
            )
        return self._pdf_to_xlsx(job)

    def _pdf_to_xlsx(self, job: ConversionJob) -> Path:
        """Extract tables from every page of a PDF and write them to an Excel workbook.

        Each page with tables produces a separate sheet ("Page 1", "Page 2", ...).
        Multiple tables on the same page are written sequentially, separated by a blank row.
        """
        output_path = job.expected_output_path
        logger.info("Extracting tables from %s", job.input_path.name)

        try:
            pdf = pdfplumber.open(job.input_path)
        except Exception as exc:
            raise EngineError(f"Failed to open PDF: {exc}") from exc

        try:
            total_pages = len(pdf.pages)
            if total_pages == 0:
                raise EngineError("PDF contains no pages.")

            wb = openpyxl.Workbook()
            # Remove the default sheet created by openpyxl; we add our own.
            default_sheet = wb.active
            wb.remove(default_sheet)

            tables_found = 0

            for page_idx, page in enumerate(pdf.pages):
                page_number = page_idx + 1
                logger.debug("Processing page %d/%d", page_number, total_pages)

                page_tables = page.extract_tables()

                if page_tables:
                    sheet_name = f"Page {page_number}"
                    ws = wb.create_sheet(title=sheet_name)
                    current_row = 1

                    for table_idx, table in enumerate(page_tables):
                        if table is None:
                            continue

                        tables_found += 1

                        # Add a header label when multiple tables exist on the same page.
                        if len(page_tables) > 1:
                            ws.cell(
                                row=current_row,
                                column=1,
                                value=f"Table {table_idx + 1}",
                            )
                            ws.cell(row=current_row, column=1).font = (
                                openpyxl.styles.Font(bold=True)
                            )
                            current_row += 1

                        for row in table:
                            for col_idx, cell_value in enumerate(row):
                                ws.cell(
                                    row=current_row,
                                    column=col_idx + 1,
                                    value=cell_value if cell_value is not None else "",
                                )
                            current_row += 1

                        # Blank separator row between tables on the same page.
                        current_row += 1

                # Update progress proportionally across pages.
                job.progress = int((page_number) / total_pages * 95)

            # Handle the case where no tables were found anywhere in the PDF.
            if tables_found == 0:
                logger.warning("No tables found in %s", job.input_path.name)
                ws = wb.create_sheet(title="Info")
                ws.cell(row=1, column=1, value="No tables detected")
                ws.cell(
                    row=2,
                    column=1,
                    value=(
                        f"pdfplumber did not find any tables in "
                        f"'{job.input_path.name}'. "
                        f"The PDF may not contain tabular data, or the tables "
                        f"may use a layout that is difficult to detect automatically."
                    ),
                )

            # Ensure at least one sheet exists (openpyxl requirement).
            if len(wb.sheetnames) == 0:
                wb.create_sheet(title="Empty")

            logger.info(
                "Extracted %d table(s) from %d page(s) in %s",
                tables_found,
                total_pages,
                job.input_path.name,
            )

            try:
                wb.save(str(output_path))
            except Exception as exc:
                raise EngineError(
                    f"Failed to write Excel file: {exc}"
                ) from exc

        finally:
            pdf.close()

        job.progress = 100
        return output_path
